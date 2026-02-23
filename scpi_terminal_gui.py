import csv
import math
import re
import threading
import subprocess
import time
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import serial
    from serial.tools import list_ports
except ImportError:  # pragma: no cover
    serial = None
    list_ports = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


XALL_GROUP_FIELDS = {
    "TEMPERATURES": [
        "Reso Temperature",
        "Vanadat Temperature",
        "Laserdiode Temperature",
        "Housing Temperature",
        "SHG Temperature",
        "THG Temperature",
        "Actual SHG Voltage",
        "Actual THG Voltage",
        "Actual SHG Current",
        "Actual THG Current",
    ],
    "STEPPER": [
        "Actual Stepper Position",
        "Actual Spot Number",
        "Spot Hours",
        "Spot Hours Remain",
        "Crystal Hours",
        "Spot Status",
        "Spot Warnings",
        "Spot Faults",
    ],
    "OTHERS": [
        "Fan Output Drive",
        "Laserdiode Current",
        "System Status Flags",
        "Scaled UV-Power",
        "Raw UV-Power",
        "Operation Hours",
        "LD Hours",
    ],
}

APP_VERSION = "v2.1"

ALL_FIELDS = [
    "Status",
    "Warnings",
    "Faults",
    "Actual Housing Temperature",
    "Actual Laserdiode Temperature",
    "Actual SHG Temperature",
    "Actual THG Temperature",
    "Operation Hours",
    "Laserdiode Hours",
    "THG Crystal Hours",
    "Actual THG Spot Hours",
    "Actual THG Spot Number",
    "Actual THG Spot Status",
    "Scaled UV-Power",
]


class SCPITerminalApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(f"SCPI Serial Terminal {APP_VERSION}")
        self.root.geometry("1240x800")

        self.ser = None
        self.reader_thread = None
        self.stop_reader = threading.Event()
        self.pause_reader = threading.Event()
        self.serial_lock = threading.Lock()

        self.response_lines: list[str] = []
        self.current_params: dict[str, str] = {}
        self.baseline_params: dict[str, str] = {}

        self.scpi_definitions: dict[str, dict[str, str]] = {}
        self.selected_response_template = ""
        self.all_xall_field_names: list[str] = []
        self.last_sent_command = ""

        self.monitor_running = False
        self.monitor_stop_event = threading.Event()
        self.monitor_thread = None
        self.monitor_csv_path: Path | None = None
        self.monitor_set_shg: float | None = None
        self.monitor_set_thg: float | None = None
        self.monitor_rows: list[dict[str, float | str]] = []

        self._build_ui()
        self._refresh_ports()
        self.root.after(100, self._announce_runtime_info)

    def _announce_runtime_info(self) -> None:
        script_path = Path(__file__).resolve()
        self._append_output(f"[INFO] App-Version: {APP_VERSION}")
        self._append_output(f"[INFO] Script: {script_path}")
        self._append_output(f"[INFO] CWD: {Path.cwd()}")

        git_root, git_branch, git_commit = self._read_git_context(script_path.parent)
        if git_root:
            self._append_output(f"[INFO] Git Root: {git_root}")
            self._append_output(f"[INFO] Git Branch: {git_branch}")
            self._append_output(f"[INFO] Git Commit: {git_commit}")
        else:
            self._append_output("[WARN] Kein Git-Repository für das gestartete Script gefunden")

    @staticmethod
    def _read_git_context(start_dir: Path) -> tuple[str, str, str]:
        try:
            root = subprocess.check_output(["git", "-C", str(start_dir), "rev-parse", "--show-toplevel"], text=True).strip()
            branch = subprocess.check_output(["git", "-C", str(start_dir), "rev-parse", "--abbrev-ref", "HEAD"], text=True).strip()
            commit = subprocess.check_output(["git", "-C", str(start_dir), "rev-parse", "--short", "HEAD"], text=True).strip()
            return root, branch, commit
        except Exception:
            return "", "", ""

    def _build_ui(self) -> None:
        top = ttk.Frame(self.root, padding=10)
        top.pack(fill=tk.X)

        ttk.Label(top, text="COM Port:").grid(row=0, column=0, sticky=tk.W)
        self.port_var = tk.StringVar()
        self.port_combo = ttk.Combobox(top, textvariable=self.port_var, width=16, state="readonly")
        self.port_combo.grid(row=0, column=1, padx=5)

        ttk.Label(top, text="Baudrate:").grid(row=0, column=2, sticky=tk.W)
        self.baud_var = tk.StringVar(value="9600")
        self.baud_combo = ttk.Combobox(top, textvariable=self.baud_var, width=10, values=["1200", "2400", "4800", "9600", "19200", "38400", "57600", "115200", "230400"], state="readonly")
        self.baud_combo.grid(row=0, column=3, padx=5)

        ttk.Label(top, text="Timeout (s):").grid(row=0, column=4, sticky=tk.W)
        self.timeout_var = tk.StringVar(value="1")
        ttk.Entry(top, textvariable=self.timeout_var, width=8).grid(row=0, column=5, padx=5)

        ttk.Button(top, text="Ports neu laden", command=self._refresh_ports).grid(row=0, column=6, padx=8)
        self.connect_btn = ttk.Button(top, text="Verbinden", command=self._toggle_connection)
        self.connect_btn.grid(row=0, column=7, padx=8)

        self.status_var = tk.StringVar(value="Nicht verbunden")
        ttk.Label(top, textvariable=self.status_var, foreground="#444").grid(row=0, column=8, sticky=tk.W, padx=4)

        self.version_var = tk.StringVar(value=f"Build: {APP_VERSION}")
        ttk.Label(top, textvariable=self.version_var, foreground="#666").grid(row=0, column=9, sticky=tk.E, padx=4)

        serial_opts = ttk.LabelFrame(self.root, text="Serielle Einstellungen", padding=10)
        serial_opts.pack(fill=tk.X, padx=10, pady=(0, 8))

        ttk.Label(serial_opts, text="Datenbits:").grid(row=0, column=0, sticky=tk.W)
        self.bytesize_var = tk.StringVar(value="8")
        ttk.Combobox(serial_opts, textvariable=self.bytesize_var, width=7, values=["5", "6", "7", "8"], state="readonly").grid(row=0, column=1, padx=(5, 12), sticky=tk.W)

        ttk.Label(serial_opts, text="Parität:").grid(row=0, column=2, sticky=tk.W)
        self.parity_var = tk.StringVar(value="None")
        ttk.Combobox(serial_opts, textvariable=self.parity_var, width=8, values=["None", "Even", "Odd", "Mark", "Space"], state="readonly").grid(row=0, column=3, padx=(5, 12), sticky=tk.W)

        ttk.Label(serial_opts, text="Stopbits:").grid(row=0, column=4, sticky=tk.W)
        self.stopbits_var = tk.StringVar(value="1")
        ttk.Combobox(serial_opts, textvariable=self.stopbits_var, width=6, values=["1", "1.5", "2"], state="readonly").grid(row=0, column=5, padx=(5, 12), sticky=tk.W)

        ttk.Label(serial_opts, text="Encoding:").grid(row=0, column=6, sticky=tk.W)
        self.encoding_var = tk.StringVar(value="ascii")
        ttk.Combobox(serial_opts, textvariable=self.encoding_var, width=8, values=["ascii", "utf-8", "latin-1"], state="readonly").grid(row=0, column=7, padx=(5, 12), sticky=tk.W)

        ttk.Label(serial_opts, text="Zeilenende:").grid(row=0, column=8, sticky=tk.W)
        self.newline_var = tk.StringVar(value="CR")
        ttk.Combobox(serial_opts, textvariable=self.newline_var, width=7, values=["CR", "LF", "CRLF", "None"], state="readonly").grid(row=0, column=9, padx=(5, 12), sticky=tk.W)

        ttk.Button(serial_opts, text="Preset 115200 8N1 CR", command=lambda: self._apply_serial_preset("115200", "8", "None", "1", "ascii", "CR")).grid(row=0, column=10, sticky=tk.W)
        ttk.Button(serial_opts, text="Preset 115200 8N1 CRLF", command=lambda: self._apply_serial_preset("115200", "8", "None", "1", "ascii", "CRLF")).grid(row=0, column=11, sticky=tk.W, padx=(8, 0))

        catalog = ttk.LabelFrame(self.root, text="SCPI Katalog", padding=10)
        catalog.pack(fill=tk.X, padx=10, pady=(0, 8))

        ttk.Button(catalog, text="SCPI Katalog laden", command=self.load_scpi_catalog).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(catalog, text="ALL/XALL Feldliste laden", command=self.load_all_xall_field_map).grid(row=0, column=1, padx=(0, 8))

        ttk.Label(catalog, text="Befehl:").grid(row=0, column=2, sticky=tk.W)
        self.catalog_cmd_var = tk.StringVar()
        self.catalog_cmd_combo = ttk.Combobox(catalog, textvariable=self.catalog_cmd_var, state="readonly", width=50)
        self.catalog_cmd_combo.grid(row=0, column=3, padx=6, sticky=tk.W)
        self.catalog_cmd_combo.bind("<<ComboboxSelected>>", lambda _e: self._on_catalog_command_selected())
        ttk.Button(catalog, text="Befehl übernehmen", command=self.use_selected_catalog_command).grid(row=0, column=4, padx=6)

        self.template_var = tk.StringVar(value="Response-Template: -")
        self.command_info_var = tk.StringVar(value="Kategorie/Info: -")
        self.field_map_var = tk.StringVar(value="ALL/XALL Feldliste: nicht geladen")
        ttk.Label(catalog, textvariable=self.template_var).grid(row=1, column=0, columnspan=5, sticky=tk.W, pady=(6, 0))
        ttk.Label(catalog, textvariable=self.command_info_var).grid(row=2, column=0, columnspan=5, sticky=tk.W)
        ttk.Label(catalog, textvariable=self.field_map_var).grid(row=3, column=0, columnspan=5, sticky=tk.W)

        cmd_frame = ttk.LabelFrame(self.root, text="SCPI Terminal", padding=10)
        cmd_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 8))

        cmd_row = ttk.Frame(cmd_frame)
        cmd_row.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(cmd_row, text="Command:").pack(side=tk.LEFT)
        self.command_var = tk.StringVar()
        cmd_entry = ttk.Entry(cmd_row, textvariable=self.command_var)
        cmd_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
        cmd_entry.bind("<Return>", lambda _e: self.send_command())

        ttk.Button(cmd_row, text="Senden", command=self.send_command).pack(side=tk.LEFT)
        ttk.Button(cmd_row, text="*IDN?", command=lambda: self._preset_send("*IDN?")).pack(side=tk.LEFT, padx=5)
        ttk.Button(cmd_row, text="IDN Terminator-Test", command=self._idn_terminator_test).pack(side=tk.LEFT, padx=5)
        ttk.Button(cmd_row, text="ALL?", command=lambda: self._preset_send("ALL?")).pack(side=tk.LEFT, padx=5)
        ttk.Button(cmd_row, text="SERVice:XALL? TEMPeratures", command=lambda: self._preset_send("SERVice:XALL? TEMPeratures")).pack(side=tk.LEFT)

        self.output = tk.Text(cmd_frame, height=18, wrap=tk.NONE)
        self.output.pack(fill=tk.BOTH, expand=True)
        scroll_y = ttk.Scrollbar(self.output, orient=tk.VERTICAL, command=self.output.yview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.output.configure(yscrollcommand=scroll_y.set)

        bottom = ttk.Frame(self.root, padding=10)
        bottom.pack(fill=tk.X)

        ttk.Button(bottom, text="Antworten als TXT speichern", command=self.save_txt).pack(side=tk.LEFT)
        ttk.Button(bottom, text="Antworten als CSV speichern", command=self.save_csv).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom, text="Log löschen", command=self.clear_output).pack(side=tk.LEFT, padx=(0, 20))

        ttk.Button(bottom, text="Aktuelle Parameter aus letztem Kommando", command=self.extract_current_params).pack(side=tk.LEFT)
        ttk.Button(bottom, text="Referenzdatei laden", command=self.load_baseline_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom, text="Parameter vergleichen", command=self.compare_params).pack(side=tk.LEFT)
        self.monitor_btn = ttk.Button(bottom, text="monitor Temp Stage SHG/THG", command=self.toggle_temp_monitor)
        self.monitor_btn.pack(side=tk.RIGHT)

        cmp_frame = ttk.LabelFrame(self.root, text="Vergleich", padding=10)
        cmp_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.compare_tree = ttk.Treeview(cmp_frame, columns=("param", "current", "baseline", "status"), show="headings", height=8)
        for col, text, width in [("param", "Parameter", 260), ("current", "Aktuell", 280), ("baseline", "Referenz", 280), ("status", "Status", 150)]:
            self.compare_tree.heading(col, text=text)
            self.compare_tree.column(col, width=width)
        self.compare_tree.pack(fill=tk.BOTH, expand=True)

        self.compare_tree.tag_configure("ok", background="#e7f8e7")
        self.compare_tree.tag_configure("diff", background="#ffe8e8")
        self.compare_tree.tag_configure("missing", background="#fff8d9")

        monitor_frame = ttk.LabelFrame(self.root, text="Temperatur-Abweichung (Ist - Soll)", padding=10)
        monitor_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.monitor_canvas = tk.Canvas(monitor_frame, height=220, bg="white")
        self.monitor_canvas.pack(fill=tk.BOTH, expand=True)
        self.monitor_canvas.bind("<Configure>", lambda _e: self._draw_monitor_plot())

        self.monitor_info_var = tk.StringVar(value="Monitoring inaktiv")
        ttk.Label(monitor_frame, textvariable=self.monitor_info_var).pack(anchor=tk.W, pady=(8, 0))

    def _apply_serial_preset(self, baud: str, bits: str, parity: str, stopbits: str, encoding: str, newline: str) -> None:
        self.baud_var.set(baud)
        self.bytesize_var.set(bits)
        self.parity_var.set(parity)
        self.stopbits_var.set(stopbits)
        self.encoding_var.set(encoding)
        self.newline_var.set(newline)
        self._append_output(f"[INFO] Preset gesetzt: {baud}, {bits}{parity[0]}{stopbits}, {newline}, {encoding}")

    def _refresh_ports(self) -> None:
        if list_ports is None:
            self.port_combo["values"] = []
            self.status_var.set("pyserial nicht installiert")
            return

        ports = [p.device for p in list_ports.comports()]
        self.port_combo["values"] = ports
        self.port_var.set(ports[0] if ports else "")
        self.status_var.set("Portliste aktualisiert" if ports else "Keine COM Ports gefunden")

    def _toggle_connection(self) -> None:
        if self.ser and self.ser.is_open:
            self._disconnect()
        else:
            self._connect()

    def _connect(self) -> None:
        if serial is None:
            messagebox.showerror("Fehler", "pyserial ist nicht installiert.\nInstalliere: pip install pyserial")
            return

        port = self.port_var.get().strip()
        if not port:
            messagebox.showwarning("Hinweis", "Bitte COM Port auswählen")
            return

        try:
            baud = int(self.baud_var.get())
            timeout = float(self.timeout_var.get())
            bytesize_map = {"5": serial.FIVEBITS, "6": serial.SIXBITS, "7": serial.SEVENBITS, "8": serial.EIGHTBITS}
            parity_map = {
                "None": serial.PARITY_NONE,
                "Even": serial.PARITY_EVEN,
                "Odd": serial.PARITY_ODD,
                "Mark": serial.PARITY_MARK,
                "Space": serial.PARITY_SPACE,
            }
            stopbits_map = {"1": serial.STOPBITS_ONE, "1.5": serial.STOPBITS_ONE_POINT_FIVE, "2": serial.STOPBITS_TWO}
            self.ser = serial.Serial(
                port=port,
                baudrate=baud,
                timeout=timeout,
                bytesize=bytesize_map.get(self.bytesize_var.get(), serial.EIGHTBITS),
                parity=parity_map.get(self.parity_var.get(), serial.PARITY_NONE),
                stopbits=stopbits_map.get(self.stopbits_var.get(), serial.STOPBITS_ONE),
            )
        except Exception as exc:
            messagebox.showerror("Verbindungsfehler", str(exc))
            return

        self.stop_reader.clear()
        self.reader_thread = threading.Thread(target=self._reader_loop, daemon=True)
        self.reader_thread.start()
        self.connect_btn.configure(text="Trennen")
        framing = f"{self.bytesize_var.get()}{self.parity_var.get()[0]}{self.stopbits_var.get()}"
        self.status_var.set(f"Verbunden: {port} @ {baud} ({framing})")
        self._append_output(f"[INFO] Verbindung hergestellt ({port}, {baud}, {framing}, {self.newline_var.get()}, {self.encoding_var.get()})")

    def _disconnect(self) -> None:
        if self.monitor_running:
            self._stop_temp_monitor("Monitoring durch Disconnect gestoppt")

        self.stop_reader.set()
        if self.reader_thread and self.reader_thread.is_alive():
            self.reader_thread.join(timeout=1)

        if self.ser:
            try:
                self.ser.close()
            except Exception:
                pass

        self.connect_btn.configure(text="Verbinden")
        self.status_var.set("Nicht verbunden")
        self._append_output("[INFO] Verbindung getrennt")

    def _reader_loop(self) -> None:
        while not self.stop_reader.is_set() and self.ser and self.ser.is_open:
            if self.pause_reader.is_set():
                time.sleep(0.01)
                continue
            try:
                with self.serial_lock:
                    raw = self.ser.readline()
                if raw:
                    text, hex_hint = self._decode_raw_bytes(raw)
                    if text:
                        self.root.after(0, self._append_output, f"RX> {text}")
                    if hex_hint:
                        self.root.after(0, self._append_output, f"[HINT] RX HEX: {hex_hint}")
                else:
                    time.sleep(0.02)
            except Exception as exc:
                self.root.after(0, self._append_output, f"[ERROR] Lesefehler: {exc}")
                break

    def _decode_raw_bytes(self, raw: bytes) -> tuple[str, str]:
        encoding = self.encoding_var.get().strip() or "ascii"
        text = raw.decode(encoding, errors="replace").strip()
        replacement_count = text.count("�")
        bad_ratio = (replacement_count / max(len(text), 1)) if text else 0
        if bad_ratio > 0.2:
            return text, raw.hex(" ")
        return text, ""

    def _preset_send(self, cmd: str) -> None:
        self.command_var.set(cmd)
        self.send_command()

    def _idn_terminator_test(self) -> None:
        if not self.ser or not self.ser.is_open:
            messagebox.showwarning("Nicht verbunden", "Bitte zuerst verbinden")
            return

        original_newline = self.newline_var.get()
        for term in ["CR", "LF", "CRLF", "None"]:
            self.newline_var.set(term)
            self.command_var.set("*IDN?")
            self.send_command()
            time.sleep(0.15)
        self.newline_var.set(original_newline)
        self._append_output("[INFO] IDN Terminator-Test gesendet (CR/LF/CRLF/None)")

    def send_command(self) -> None:
        cmd = self.command_var.get().strip()
        if not cmd:
            return
        if not self.ser or not self.ser.is_open:
            messagebox.showwarning("Nicht verbunden", "Bitte zuerst verbinden")
            return

        try:
            newline_map = {"LF": "\n", "CR": "\r", "CRLF": "\r\n", "None": ""}
            line_end = newline_map.get(self.newline_var.get(), "\r")
            payload = (cmd + line_end).encode(self.encoding_var.get(), errors="replace")
            with self.serial_lock:
                self.ser.write(payload)
            self.last_sent_command = cmd
            self._append_output(f"TX> {cmd}")
            self._append_output(f"[DEBUG] TX HEX: {payload.hex(' ')}")
            self.command_var.set("")
        except Exception as exc:
            messagebox.showerror("Senden fehlgeschlagen", str(exc))

    def _append_output(self, line: str) -> None:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        full = f"{stamp} {line}"
        self.response_lines.append(full)
        self.output.insert(tk.END, full + "\n")
        self.output.see(tk.END)

    def clear_output(self) -> None:
        self.response_lines.clear()
        self.output.delete("1.0", tk.END)

    def save_txt(self) -> None:
        if not self.response_lines:
            messagebox.showinfo("Hinweis", "Keine Daten zum Speichern")
            return
        path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text", "*.txt")])
        if path:
            Path(path).write_text("\n".join(self.response_lines), encoding="utf-8")
            messagebox.showinfo("Gespeichert", f"TXT gespeichert:\n{path}")

    def save_csv(self) -> None:
        if not self.response_lines:
            messagebox.showinfo("Hinweis", "Keine Daten zum Speichern")
            return

        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(["timestamp", "direction", "message"])
            for line in self.response_lines:
                ts = line[:19]
                rest = line[20:] if len(line) > 20 else ""
                if rest.startswith("TX>"):
                    direction = "TX"
                    msg = rest[4:]
                elif rest.startswith("RX>"):
                    direction = "RX"
                    msg = rest[4:]
                else:
                    direction = "INFO"
                    msg = rest
                writer.writerow([ts, direction, msg])

        messagebox.showinfo("Gespeichert", f"CSV gespeichert:\n{path}")

    def load_scpi_catalog(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Katalog", "*.xlsx *.xlsm *.csv *.tsv *.txt"), ("Alle", "*.*")])
        if not path:
            return
        try:
            rows = self._load_catalog_rows(path)
            parsed = self._parse_scpi_rows(rows)
            self.scpi_definitions = parsed
            labels = sorted(parsed.keys())
            self.catalog_cmd_combo["values"] = labels
            if labels:
                self.catalog_cmd_var.set(labels[0])
                self._on_catalog_command_selected()
                messagebox.showinfo("SCPI geladen", f"{len(labels)} Befehle gefunden")
            else:
                messagebox.showwarning("Hinweis", "Keine SCPI-Befehle im Katalog erkannt")
        except Exception as exc:
            messagebox.showerror("Katalog Fehler", str(exc))

    def _load_catalog_rows(self, path: str) -> list[dict[str, str]]:
        ext = Path(path).suffix.lower()
        if ext in {".xlsx", ".xlsm"}:
            return self._load_catalog_rows_xlsx(path)
        return self._load_catalog_rows_delimited(path)

    def _load_catalog_rows_xlsx(self, path: str) -> list[dict[str, str]]:
        if load_workbook is None:
            raise RuntimeError("openpyxl fehlt. Bitte installieren: pip install openpyxl")

        wb = load_workbook(path, data_only=True, read_only=True)
        rows: list[dict[str, str]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_rows = list(ws.iter_rows(min_row=1, max_row=6000, values_only=True))
            if not raw_rows:
                continue
            header_idx, header = self._detect_header(raw_rows)
            normalized = [self._norm_key(h) for h in header]
            for row in raw_rows[header_idx + 1 :]:
                values = ["" if v is None else str(v).strip() for v in row]
                item = {normalized[i]: values[i] if i < len(values) else "" for i in range(len(normalized))}
                if any(v for v in item.values()):
                    rows.append(item)
        return rows

    def _load_catalog_rows_delimited(self, path: str) -> list[dict[str, str]]:
        text = Path(path).read_text(encoding="utf-8", errors="replace")
        sample = text.splitlines()[0] if text.splitlines() else ""
        delimiter = "\t" if "\t" in sample else ";" if ";" in sample else ","
        reader = csv.DictReader(text.splitlines(), delimiter=delimiter)
        rows = []
        for row in reader:
            normalized = {self._norm_key(k): ("" if v is None else str(v).strip()) for k, v in row.items() if k}
            rows.append(normalized)
        return rows

    def _parse_scpi_rows(self, rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
        result: dict[str, dict[str, str]] = {}
        for row in rows:
            command = self._pick(row, ["scpi command", "command", "befehl", "scpi", "cmd"])
            if not command:
                continue

            category = self._pick(row, ["category", "kategorie", "group"]) or "-"
            response = self._pick(row, ["response", "antwort", "resp", "rückgabe"])
            instance = self._pick(row, ["instance"])
            parameter = self._pick(row, ["parameter"])
            unit = self._pick(row, ["unit", "einheit"])
            description = self._pick(row, ["description", "beschreibung"]) or ""

            label = f"{category} | {command}"
            result[label] = {
                "category": category,
                "command": command,
                "response": response,
                "instance": instance,
                "parameter": parameter,
                "unit": unit,
                "description": description,
            }
        return result

    def load_all_xall_field_map(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Feldliste", "*.txt *.csv"), ("Alle", "*.*")])
        if not path:
            return
        content = Path(path).read_text(encoding="utf-8", errors="replace").splitlines()
        fields = []
        for line in content:
            token = line.strip().strip(";")
            if token and token.lower() not in {"field", "parameter", "name"}:
                fields.append(token)
        self.all_xall_field_names = fields
        self.field_map_var.set(f"ALL/XALL Feldliste: {len(fields)} Einträge geladen")

    def _on_catalog_command_selected(self) -> None:
        selected = self.catalog_cmd_var.get().strip()
        entry = self.scpi_definitions.get(selected, {})
        response = entry.get("response", "")
        self.selected_response_template = response
        self.template_var.set(f"Response-Template: {response if response else '-'}")
        info = f"Kategorie: {entry.get('category', '-')} | Instance: {entry.get('instance', '-')} | Unit: {entry.get('unit', '-') }"
        self.command_info_var.set(info)

    def use_selected_catalog_command(self) -> None:
        selected = self.catalog_cmd_var.get().strip()
        if not selected or selected not in self.scpi_definitions:
            messagebox.showwarning("Hinweis", "Bitte zuerst einen Katalog-Befehl auswählen")
            return
        self.command_var.set(self.scpi_definitions[selected]["command"])

    @staticmethod
    def _norm_key(value: str) -> str:
        return str(value).strip().lower() if value else ""

    def _pick(self, row: dict[str, str], variants: list[str]) -> str:
        for key in variants:
            norm = self._norm_key(key)
            if norm in row and row[norm].strip():
                return row[norm].strip()
        return ""

    @staticmethod
    def _detect_header(rows: list[tuple]) -> tuple[int, list[str]]:
        best_idx = 0
        best_score = -1
        best_header: list[str] = []
        keys = ("category", "command", "befehl", "parameter", "response", "antwort", "unit", "description")
        for idx, row in enumerate(rows[:25]):
            header = ["" if v is None else str(v).strip().lower() for v in row]
            score = sum(1 for c in header if any(k in c for k in keys))
            if score > best_score:
                best_idx, best_score, best_header = idx, score, header
        return best_idx, best_header

    @staticmethod
    def _strip_prefix(line: str) -> str:
        trimmed = line[20:] if len(line) > 20 else line
        if trimmed.startswith("RX>"):
            return trimmed[4:].strip()
        if trimmed.startswith("TX>"):
            return ""
        return trimmed.strip()

    def _extract_last_response_block(self) -> list[str]:
        tx_indexes = [i for i, l in enumerate(self.response_lines) if " TX>" in l]
        if not tx_indexes:
            return [self._strip_prefix(l) for l in self.response_lines if self._strip_prefix(l)]

        start = tx_indexes[-1]
        result = []
        for line in self.response_lines[start + 1 :]:
            if " TX>" in line:
                break
            stripped = self._strip_prefix(line)
            if stripped:
                result.append(stripped)
        return result

    @staticmethod
    def _parse_key_values(lines: list[str]) -> dict[str, str]:
        result: dict[str, str] = {}
        for raw in lines:
            line = raw.strip()
            for segment in re.split(r"[;|]", line):
                segment = segment.strip()
                if not segment:
                    continue
                for sep in ("=", ":", ","):
                    if sep in segment:
                        left, right = segment.split(sep, 1)
                        key = left.strip()
                        value = right.strip()
                        if key:
                            result[key] = value
                        break
        return result

    @staticmethod
    def _extract_template_fields(template: str) -> list[str]:
        if not template:
            return []
        fields = re.findall(r"[A-Za-z][A-Za-z0-9_\-/]*", template)
        return [f for f in fields if len(f) > 1][:250]

    @staticmethod
    def _split_values(lines: list[str]) -> list[str]:
        values = []
        for line in lines:
            for token in re.split(r"[,;]", line):
                token = token.strip()
                if token:
                    values.append(token)
        return values

    def _xall_fields_from_command(self, command_text: str) -> list[str]:
        upper = command_text.upper().strip()
        if "XALL?" not in upper:
            return []
        for group, fields in XALL_GROUP_FIELDS.items():
            if group in upper:
                return fields
        return []

    def _parse_all_xall_response(self, lines: list[str], command_text: str) -> dict[str, str]:
        values = self._split_values(lines)
        if not values:
            return {}

        upper = command_text.upper().strip()
        fields: list[str] = []
        if upper.endswith("ALL?"):
            fields = ALL_FIELDS
        elif "XALL?" in upper:
            fields = self._xall_fields_from_command(command_text)

        if not fields and self.all_xall_field_names:
            fields = self.all_xall_field_names

        if not fields:
            return {}

        parsed: dict[str, str] = {}
        for idx, field in enumerate(fields):
            parsed[field] = values[idx] if idx < len(values) else ""
        return parsed

    def _parse_response_by_template(self, lines: list[str], template: str) -> dict[str, str]:
        command_text = self.last_sent_command.strip().upper()
        if command_text.endswith("ALL?") or "XALL?" in command_text:
            parsed = self._parse_all_xall_response(lines, command_text)
            if parsed:
                return parsed

        keys = self._extract_template_fields(template)
        if not keys:
            return self._parse_key_values(lines)

        values = self._split_values(lines)
        if len(values) < len(keys):
            return self._parse_key_values(lines)

        return {key: values[i] for i, key in enumerate(keys)}

    def extract_current_params(self) -> None:
        response_block = self._extract_last_response_block()
        if not response_block:
            messagebox.showwarning("Hinweis", "Keine Antwortdaten gefunden")
            return
        parsed = self._parse_response_by_template(response_block, self.selected_response_template)
        if not parsed:
            parsed = {"RAW_RESPONSE": " | ".join(response_block)}
        self.current_params = parsed
        messagebox.showinfo("Parameter", f"Aktuelle Parameter erkannt: {len(parsed)}")

    def load_baseline_file(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Vergleichsdateien", "*.txt *.csv *.xlsx"), ("Alle", "*.*")])
        if not path:
            return

        try:
            ext = Path(path).suffix.lower()
            parsed: dict[str, str]
            if ext == ".csv":
                parsed = self._parse_csv_file(path)
            elif ext in {".xlsx", ".xlsm"}:
                parsed = self._parse_excel_baseline(path)
            else:
                content = Path(path).read_text(encoding="utf-8", errors="replace").splitlines()
                parsed = self._parse_key_values(content)

            if not parsed:
                parsed = {"RAW_RESPONSE": Path(path).read_text(encoding="utf-8", errors="replace").strip()}

            self.baseline_params = parsed
            messagebox.showinfo("Referenz geladen", f"Datei: {path}\nParameter: {len(parsed)}")
        except Exception as exc:
            messagebox.showerror("Fehler", str(exc))

    @staticmethod
    def _parse_csv_file(path: str) -> dict[str, str]:
        result: dict[str, str] = {}
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=";")
            for row in reader:
                if len(row) >= 2 and row[0].strip().lower() not in {"timestamp", "direction", "message"}:
                    key = row[0].strip()
                    value = row[1].strip() if len(row) > 1 else ""
                    if key:
                        result[key] = value
        return result

    def _parse_excel_baseline(self, path: str) -> dict[str, str]:
        if load_workbook is None:
            raise RuntimeError("openpyxl fehlt für Excel Vergleichsdateien")
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        lines = []
        for row in ws.iter_rows(min_row=1, max_row=600, values_only=True):
            vals = ["" if x is None else str(x).strip() for x in row if str(x).strip()]
            if vals:
                lines.append("=".join(vals[:2]) if len(vals) >= 2 else vals[0])
        return self._parse_key_values(lines)

    def compare_params(self) -> None:
        if not self.current_params:
            messagebox.showwarning("Hinweis", "Bitte zuerst aktuelle Parameter aus letzter Antwort extrahieren")
            return
        if not self.baseline_params:
            messagebox.showwarning("Hinweis", "Bitte zuerst eine Referenzdatei laden")
            return

        for item in self.compare_tree.get_children():
            self.compare_tree.delete(item)

        keys = sorted(set(self.current_params) | set(self.baseline_params))
        diff_count = 0
        for key in keys:
            current = self.current_params.get(key)
            baseline = self.baseline_params.get(key)

            if current is None:
                status, tag = "Fehlt aktuell", "missing"
                diff_count += 1
            elif baseline is None:
                status, tag = "Fehlt Referenz", "missing"
                diff_count += 1
            elif current == baseline:
                status, tag = "OK", "ok"
            else:
                status, tag = "Abweichung", "diff"
                diff_count += 1

            self.compare_tree.insert("", tk.END, values=(key, current or "", baseline or "", status), tags=(tag,))

        self.status_var.set(f"Vergleich abgeschlossen: {diff_count} Abweichung(en)")

    def toggle_temp_monitor(self) -> None:
        if self.monitor_running:
            self._stop_temp_monitor("Monitoring gestoppt")
            return
        self._start_temp_monitor()

    def _start_temp_monitor(self) -> None:
        if not self.ser or not self.ser.is_open:
            messagebox.showwarning("Nicht verbunden", "Bitte zuerst verbinden")
            return

        path = filedialog.asksaveasfilename(
            title="CSV für Temperatur-Monitoring speichern",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=f"temp_monitor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not path:
            return

        try:
            set_shg = self._query_temperature_value([
                "SOURce:TEMPerature:LEVel:SET? TEMP_STAGE_SHG",
                "SOURce:TEMPerature:LEVel:SET? TEMP_STAGE SHG",
            ])
            set_thg = self._query_temperature_value([
                "SOURce:TEMPerature:LEVel:SET? TEMP_STAGE_THG",
            ])
        except RuntimeError as exc:
            messagebox.showerror("Monitoring", str(exc))
            return

        self.monitor_set_shg = set_shg
        self.monitor_set_thg = set_thg
        self.monitor_csv_path = Path(path)
        self.monitor_rows.clear()
        self.monitor_stop_event.clear()
        self.monitor_running = True
        self.monitor_btn.configure(text="Monitoring stoppen")
        self.monitor_info_var.set(f"Monitoring aktiv | Set SHG={set_shg:.3f}, Set THG={set_thg:.3f}")
        self._append_output(f"[INFO] Monitoring gestartet, CSV: {self.monitor_csv_path}")

        with self.monitor_csv_path.open("w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile, delimiter=";")
            writer.writerow(["timestamp", "elapsed_s", "set_shg", "actual_shg", "delta_shg", "set_thg", "actual_thg", "delta_thg"])

        self.monitor_thread = threading.Thread(target=self._monitor_temp_loop, daemon=True)
        self.monitor_thread.start()

    def _stop_temp_monitor(self, reason: str) -> None:
        self.monitor_stop_event.set()
        self.monitor_running = False
        self.monitor_btn.configure(text="monitor Temp Stage SHG/THG")
        self.monitor_info_var.set(reason)
        self._append_output(f"[INFO] {reason}")

    def _monitor_temp_loop(self) -> None:
        start = time.time()
        while not self.monitor_stop_event.is_set():
            try:
                actual_shg = self._query_temperature_value(["SOURce:TEMPerature:ACTual? TEMP_STAGE_SHG"])
                actual_thg = self._query_temperature_value(["SOURce:TEMPerature:ACTual? TEMP_STAGE_THG"])
                elapsed = time.time() - start
                stamp = datetime.now().isoformat(timespec="milliseconds")

                delta_shg = actual_shg - (self.monitor_set_shg if self.monitor_set_shg is not None else 0.0)
                delta_thg = actual_thg - (self.monitor_set_thg if self.monitor_set_thg is not None else 0.0)

                row = {
                    "timestamp": stamp,
                    "elapsed_s": elapsed,
                    "set_shg": self.monitor_set_shg if self.monitor_set_shg is not None else math.nan,
                    "actual_shg": actual_shg,
                    "delta_shg": delta_shg,
                    "set_thg": self.monitor_set_thg if self.monitor_set_thg is not None else math.nan,
                    "actual_thg": actual_thg,
                    "delta_thg": delta_thg,
                }
                self.monitor_rows.append(row)

                if self.monitor_csv_path:
                    with self.monitor_csv_path.open("a", newline="", encoding="utf-8") as csvfile:
                        writer = csv.writer(csvfile, delimiter=";")
                        writer.writerow([
                            row["timestamp"],
                            f"{row['elapsed_s']:.3f}",
                            f"{row['set_shg']:.6f}",
                            f"{row['actual_shg']:.6f}",
                            f"{row['delta_shg']:.6f}",
                            f"{row['set_thg']:.6f}",
                            f"{row['actual_thg']:.6f}",
                            f"{row['delta_thg']:.6f}",
                        ])

                self.root.after(0, self._draw_monitor_plot)
                self.root.after(0, self.monitor_info_var.set, f"Monitoring aktiv | Samples: {len(self.monitor_rows)} | ΔSHG={delta_shg:.3f} | ΔTHG={delta_thg:.3f}")
            except Exception as exc:
                self.root.after(0, self._stop_temp_monitor, f"Monitoring Fehler: {exc}")
                return

            time.sleep(0.1)

        self.root.after(0, self._draw_monitor_plot)

    def _query_temperature_value(self, commands: list[str]) -> float:
        if not self.ser or not self.ser.is_open:
            raise RuntimeError("Serielle Verbindung nicht verfügbar")

        errors: list[str] = []
        for cmd in commands:
            response = self._query_scpi(cmd)
            value = self._extract_first_float(response)
            if value is not None:
                return value
            errors.append(f"{cmd} -> '{response}'")

        raise RuntimeError("Temperaturantwort konnte nicht geparst werden: " + " | ".join(errors))

    def _query_scpi(self, cmd: str) -> str:
        if not self.ser or not self.ser.is_open:
            raise RuntimeError("Nicht verbunden")

        newline_map = {"LF": "\n", "CR": "\r", "CRLF": "\r\n", "None": ""}
        line_end = newline_map.get(self.newline_var.get(), "\r")
        payload = (cmd + line_end).encode(self.encoding_var.get(), errors="replace")

        self.pause_reader.set()
        try:
            with self.serial_lock:
                self.ser.reset_input_buffer()
                self.ser.write(payload)
                self.ser.flush()
                end_time = time.time() + max(float(self.timeout_var.get() or "1"), 0.2) + 0.8
                while time.time() < end_time:
                    raw = self.ser.readline()
                    if not raw:
                        continue
                    decoded, _ = self._decode_raw_bytes(raw)
                    if decoded:
                        self.root.after(0, self._append_output, f"TX> {cmd}")
                        self.root.after(0, self._append_output, f"RX> {decoded}")
                        return decoded
        finally:
            self.pause_reader.clear()

        raise RuntimeError(f"Keine Antwort für Kommando: {cmd}")

    @staticmethod
    def _extract_first_float(text: str) -> float | None:
        match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group(0))
        except ValueError:
            return None

    def _draw_monitor_plot(self) -> None:
        canvas = self.monitor_canvas
        canvas.delete("all")
        width = canvas.winfo_width()
        height = canvas.winfo_height()
        if width < 40 or height < 40:
            return

        left, top, right, bottom = 40, 10, width - 10, height - 30
        canvas.create_rectangle(left, top, right, bottom, outline="#bbbbbb")

        rows = self.monitor_rows[-300:]
        if len(rows) < 2:
            canvas.create_text(width / 2, height / 2, text="Noch keine ausreichenden Daten", fill="#888")
            return

        deltas_shg = [float(r["delta_shg"]) for r in rows]
        deltas_thg = [float(r["delta_thg"]) for r in rows]
        all_vals = deltas_shg + deltas_thg
        min_v, max_v = min(all_vals), max(all_vals)
        if math.isclose(min_v, max_v, abs_tol=1e-12):
            min_v -= 1.0
            max_v += 1.0

        x_step = (right - left) / max(len(rows) - 1, 1)

        def to_xy(idx: int, val: float) -> tuple[float, float]:
            x = left + idx * x_step
            y = bottom - ((val - min_v) / (max_v - min_v)) * (bottom - top)
            return x, y

        shg_points = [to_xy(i, v) for i, v in enumerate(deltas_shg)]
        thg_points = [to_xy(i, v) for i, v in enumerate(deltas_thg)]

        canvas.create_line(*[coord for p in shg_points for coord in p], fill="#0077cc", width=2)
        canvas.create_line(*[coord for p in thg_points for coord in p], fill="#cc5500", width=2)

        zero_y = bottom - ((0 - min_v) / (max_v - min_v)) * (bottom - top)
        if top <= zero_y <= bottom:
            canvas.create_line(left, zero_y, right, zero_y, fill="#999999", dash=(3, 3))

        canvas.create_text(left + 50, top + 10, text="SHG Δ", fill="#0077cc", anchor="w")
        canvas.create_text(left + 130, top + 10, text="THG Δ", fill="#cc5500", anchor="w")
        canvas.create_text(5, top, text=f"{max_v:.2f}", anchor="w", fill="#666")
        canvas.create_text(5, bottom, text=f"{min_v:.2f}", anchor="w", fill="#666")


def main() -> None:
    root = tk.Tk()
    app = SCPITerminalApp(root)

    def on_close() -> None:
        app._disconnect()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
